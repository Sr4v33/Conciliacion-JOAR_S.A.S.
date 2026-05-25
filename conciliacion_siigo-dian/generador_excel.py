"""
Construcción del reporte Excel de conciliación
Rectificadora JOAR S.A.S.
"""

from __future__ import annotations

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from conciliador import ResultadoConciliacion, RegistroFaltante
from utils import logger, formatear_moneda


# ---------------------------------------------------------------------------
# Paleta de colores
# ---------------------------------------------------------------------------

COLOR_AZUL_OSCURO  = "1F3864"   # Encabezados de tabla
COLOR_AZUL_MEDIO   = "2E75B6"   # Sub-encabezados
COLOR_VERDE        = "1E8449"   # Métrica positiva (conciliadas)
COLOR_ROJO         = "C0392B"   # Métrica negativa (faltantes) / encabezado faltantes
COLOR_AMARILLO_CLARO = "FFF2CC" # Filas alternas
COLOR_GRIS_CLARO   = "F2F2F2"   # Filas alternas alternadas
COLOR_NARANJA      = "F0A500"   # Prioridad Alta
COLOR_ROJO_CLARO   = "FADBD8"   # Prioridad Crítica (fondo fila)
COLOR_BLANCO       = "FFFFFF"
COLOR_ENCABEZADO_FALTANTES = "C0392B"

EMPRESA = "Rectificadora JOAR S.A.S."
NIT = "NIT 901363706-7"


# ---------------------------------------------------------------------------
# Estilos reutilizables
# ---------------------------------------------------------------------------

def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _font(bold=False, color="000000", size=11, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")


def _border_thin() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _align(horizontal="left", vertical="center", wrap=False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


# ---------------------------------------------------------------------------
# Generador principal
# ---------------------------------------------------------------------------

def generar_reporte(resultado: ResultadoConciliacion, carpeta_salida: str) -> str:
    """
    Genera el archivo Excel de reporte con las hojas Resumen y Faltantes.
    Retorna la ruta completa del archivo generado.
    """
    os.makedirs(carpeta_salida, exist_ok=True)

    # Nombre del archivo: Conciliacion_Siigo_DIAN_YYYYMM.xlsx
    ahora = datetime.now()
    nombre_archivo = f"Conciliacion_Siigo_DIAN_{ahora.strftime('%Y%m')}.xlsx"
    ruta_salida = os.path.join(carpeta_salida, nombre_archivo)

    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws_resumen = wb.create_sheet("Resumen")
    ws_faltantes = wb.create_sheet("Faltantes")

    _construir_hoja_resumen(ws_resumen, resultado)
    _construir_hoja_faltantes(ws_faltantes, resultado)

    wb.save(ruta_salida)
    logger.info("Reporte Excel generado: %s", ruta_salida)
    return ruta_salida


# ---------------------------------------------------------------------------
# Hoja 1: Resumen
# ---------------------------------------------------------------------------

def _construir_hoja_resumen(ws, resultado: ResultadoConciliacion):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18

    fila = 1

    # --- Encabezado corporativo ---
    fila = _encabezado_corporativo(ws, fila, "REPORTE DE CONCILIACIÓN DIAN–SIIGO", COLOR_AZUL_OSCURO)
    fila += 1

    # --- Información del período ---
    fila = _seccion_titulo(ws, fila, "INFORMACIÓN DEL PERÍODO", COLOR_AZUL_MEDIO)
    fila = _fila_dato(ws, fila, "Fecha y hora de ejecución", resultado.timestamp_ejecucion)
    fila = _fila_dato(ws, fila, "Período conciliado", resultado.periodo_etiqueta)
    fila = _fila_dato(
        ws, fila, "Rango de fechas",
        f"{resultado.periodo_inicio} – {resultado.periodo_fin}",
    )
    fila += 1

    # --- Métricas de conciliación ---
    fila = _seccion_titulo(ws, fila, "MÉTRICAS DE CONCILIACIÓN", COLOR_AZUL_MEDIO)
    fila = _fila_metrica(ws, fila, "Total facturas reportadas por DIAN",
                         str(resultado.total_dian), COLOR_AZUL_OSCURO)
    fila = _fila_metrica(ws, fila, "Total facturas causadas en Siigo",
                         str(resultado.total_siigo), COLOR_AZUL_OSCURO)
    fila = _fila_metrica(ws, fila, "Total facturas conciliadas ✔",
                         str(resultado.total_conciliadas), COLOR_VERDE)
    fila = _fila_metrica(ws, fila, "Facturas faltantes en DIAN ✘",
                         str(resultado.total_faltantes), COLOR_ROJO)
    fila = _fila_metrica(ws, fila, "Valor acumulado faltantes en DIAN",
                         formatear_moneda(resultado.valor_faltantes), COLOR_ROJO)
    fila += 1

    # --- Estado general ---
    fila = _seccion_titulo(ws, fila, "ESTADO GENERAL", COLOR_AZUL_MEDIO)
    if resultado.total_faltantes == 0:
        mensaje = "✔  Conciliación exitosa. Todos los registros de Siigo tienen correspondencia en DIAN."
        color_estado = COLOR_VERDE
    else:
        mensaje = (
            f"⚠  Se encontraron {resultado.total_faltantes} factura(s) causadas en Siigo "
            f"sin validación en DIAN por un valor total de "
            f"{formatear_moneda(resultado.valor_faltantes)}. "
            "Revisar hoja 'Faltantes' para el detalle."
        )
        color_estado = COLOR_ROJO

    ws.merge_cells(f"A{fila}:C{fila}")
    celda = ws[f"A{fila}"]
    celda.value = mensaje
    celda.fill = _fill(color_estado)
    celda.font = _font(bold=True, color=COLOR_BLANCO, size=11)
    celda.alignment = _align("left", "center", wrap=True)
    celda.border = _border_thin()
    ws.row_dimensions[fila].height = 40
    fila += 2

    # --- Pie de página ---
    ws.merge_cells(f"A{fila}:C{fila}")
    pie = ws[f"A{fila}"]
    pie.value = f"Robot RPA — Conciliación DIAN-Siigo | {EMPRESA} | Ejecutado el {resultado.timestamp_ejecucion}"
    pie.font = _font(italic=True, color="7F7F7F", size=9)
    pie.alignment = _align("center")


def _encabezado_corporativo(ws, fila: int, titulo: str, color: str) -> int:
    # Fila empresa
    ws.merge_cells(f"A{fila}:C{fila}")
    c = ws[f"A{fila}"]
    c.value = EMPRESA
    c.fill = _fill(color)
    c.font = _font(bold=True, color=COLOR_BLANCO, size=14)
    c.alignment = _align("center", "center")
    c.border = _border_thin()
    ws.row_dimensions[fila].height = 28
    fila += 1

    # Fila NIT
    ws.merge_cells(f"A{fila}:C{fila}")
    c = ws[f"A{fila}"]
    c.value = NIT
    c.fill = _fill(color)
    c.font = _font(color=COLOR_BLANCO, size=10, italic=True)
    c.alignment = _align("center", "center")
    c.border = _border_thin()
    ws.row_dimensions[fila].height = 18
    fila += 1

    # Fila título del reporte
    ws.merge_cells(f"A{fila}:C{fila}")
    c = ws[f"A{fila}"]
    c.value = titulo
    c.fill = _fill(COLOR_AZUL_MEDIO)
    c.font = _font(bold=True, color=COLOR_BLANCO, size=12)
    c.alignment = _align("center", "center")
    c.border = _border_thin()
    ws.row_dimensions[fila].height = 22
    fila += 1

    return fila


def _seccion_titulo(ws, fila: int, titulo: str, color: str) -> int:
    ws.merge_cells(f"A{fila}:C{fila}")
    c = ws[f"A{fila}"]
    c.value = titulo
    c.fill = _fill(color)
    c.font = _font(bold=True, color=COLOR_BLANCO, size=10)
    c.alignment = _align("left", "center")
    c.border = _border_thin()
    ws.row_dimensions[fila].height = 18
    return fila + 1


def _fila_dato(ws, fila: int, etiqueta: str, valor: str) -> int:
    c_etiq = ws[f"A{fila}"]
    c_etiq.value = etiqueta
    c_etiq.fill = _fill(COLOR_GRIS_CLARO)
    c_etiq.font = _font(bold=True, size=10)
    c_etiq.alignment = _align("left", "center")
    c_etiq.border = _border_thin()

    ws.merge_cells(f"B{fila}:C{fila}")
    c_val = ws[f"B{fila}"]
    c_val.value = valor
    c_val.fill = _fill(COLOR_BLANCO)
    c_val.font = _font(size=10)
    c_val.alignment = _align("left", "center")
    c_val.border = _border_thin()

    ws.row_dimensions[fila].height = 18
    return fila + 1


def _fila_metrica(ws, fila: int, etiqueta: str, valor: str, color_valor: str) -> int:
    c_etiq = ws[f"A{fila}"]
    c_etiq.value = etiqueta
    c_etiq.fill = _fill(COLOR_GRIS_CLARO if fila % 2 == 0 else COLOR_AMARILLO_CLARO)
    c_etiq.font = _font(bold=True, size=10)
    c_etiq.alignment = _align("left", "center")
    c_etiq.border = _border_thin()

    ws.merge_cells(f"B{fila}:C{fila}")
    c_val = ws[f"B{fila}"]
    c_val.value = valor
    c_val.fill = _fill(color_valor)
    c_val.font = _font(bold=True, color=COLOR_BLANCO, size=12)
    c_val.alignment = _align("center", "center")
    c_val.border = _border_thin()

    ws.row_dimensions[fila].height = 22
    return fila + 1


# ---------------------------------------------------------------------------
# Hoja 2: Faltantes
# ---------------------------------------------------------------------------

HEADERS_FALTANTES = [
    "Comprobante", "Fecha Elaboración", "NIT Proveedor", "Razón Social",
    "Factura Proveedor", "Base Gravable", "Base Exenta", "IVA", "Total",
]

ANCHOS_FALTANTES = [18, 18, 16, 36, 18, 16, 16, 14, 16, 14, 12]


def _color_prioridad(prioridad: str) -> str:
    if prioridad == "Crítica":
        return "FADBD8"   # rojo claro
    elif prioridad == "Alta":
        return "FDEBD0"   # naranja claro
    return "FFFDE7"       # amarillo claro (Media)


def _construir_hoja_faltantes(ws, resultado: ResultadoConciliacion):
    ws.sheet_view.showGridLines = False

    fila = 1

    # --- Encabezado corporativo ---
    ncols = len(HEADERS_FALTANTES)
    ultima_col = get_column_letter(ncols)

    fila = _encabezado_faltantes(ws, fila, ncols, resultado)
    fila += 1

    # --- Headers de la tabla ---
    for col_idx, header in enumerate(HEADERS_FALTANTES, start=1):
        c = ws.cell(row=fila, column=col_idx, value=header)
        c.fill = _fill(COLOR_AZUL_OSCURO)
        c.font = _font(bold=True, color=COLOR_BLANCO, size=10)
        c.alignment = _align("center", "center", wrap=True)
        c.border = _border_thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = ANCHOS_FALTANTES[col_idx - 1]
    ws.row_dimensions[fila].height = 22
    fila += 1

    # --- Filas de datos ---
    fila_inicio_datos = fila

    if not resultado.faltantes:
        ws.merge_cells(f"A{fila}:{ultima_col}{fila}")
        c = ws[f"A{fila}"]
        c.value = "✔  No se encontraron facturas faltantes en DIAN para este período."
        c.fill = _fill(COLOR_VERDE)
        c.font = _font(bold=True, color=COLOR_BLANCO, size=11)
        c.alignment = _align("center", "center")
        fila += 1
    else:
        for reg in resultado.faltantes:
            color_fila = _color_prioridad(reg.prioridad)
            valores = [
                reg.comprobante,
                reg.fecha_elaboracion,
                reg.identificacion,
                reg.razon_social,
                reg.factura_proveedor,
                reg.base_gravable,
                reg.base_exenta,
                reg.iva,
                reg.total,
            ]
            for col_idx, val in enumerate(valores, start=1):
                c = ws.cell(row=fila, column=col_idx, value=val)
                c.fill = _fill(color_fila)
                c.font = _font(size=10)
                c.border = _border_thin()
                # Alinear números a la derecha
                if col_idx in (6, 7, 8, 9, 10):
                    c.alignment = _align("right", "center")
                    if col_idx in (6, 7, 8, 9):
                        c.number_format = '#,##0.00'
                else:
                    c.alignment = _align("left", "center")
            ws.row_dimensions[fila].height = 16
            fila += 1

    # --- Fila totalizadora ---
    fila += 0  # sin espacio extra
    ws.merge_cells(f"A{fila}:H{fila}")
    c_total_label = ws[f"A{fila}"]
    c_total_label.value = f"TOTAL ({resultado.total_faltantes} factura(s) faltante(s))"
    c_total_label.fill = _fill(COLOR_ROJO)
    c_total_label.font = _font(bold=True, color=COLOR_BLANCO, size=11)
    c_total_label.alignment = _align("right", "center")
    c_total_label.border = _border_thin()

    c_total_val = ws.cell(row=fila, column=9, value=resultado.valor_faltantes)
    c_total_val.fill = _fill(COLOR_ROJO)
    c_total_val.font = _font(bold=True, color=COLOR_BLANCO, size=11)
    c_total_val.alignment = _align("right", "center")
    c_total_val.border = _border_thin()
    c_total_val.number_format = '#,##0.00'

    ws.row_dimensions[fila].height = 22
    fila += 2

    # --- Pie ---
    ws.merge_cells(f"A{fila}:{ultima_col}{fila}")
    pie = ws[f"A{fila}"]
    pie.value = f"Robot RPA — Conciliación DIAN-Siigo | {EMPRESA} | Ejecutado el {resultado.timestamp_ejecucion}"
    pie.font = _font(italic=True, color="7F7F7F", size=9)
    pie.alignment = _align("center")

    # Freeze panes debajo del encabezado de tabla
    ws.freeze_panes = f"A{fila_inicio_datos}"


def _encabezado_faltantes(ws, fila: int, ncols: int, resultado: ResultadoConciliacion) -> int:
    ultima_col = get_column_letter(ncols)

    def fila_enc(texto, color_fondo, bold=True, size=12, italic=False):
        nonlocal fila
        ws.merge_cells(f"A{fila}:{ultima_col}{fila}")
        c = ws[f"A{fila}"]
        c.value = texto
        c.fill = _fill(color_fondo)
        c.font = _font(bold=bold, color=COLOR_BLANCO, size=size, italic=italic)
        c.alignment = _align("center", "center")
        c.border = _border_thin()
        ws.row_dimensions[fila].height = 24 if bold else 18
        fila += 1

    fila_enc(EMPRESA, COLOR_ENCABEZADO_FALTANTES, bold=True, size=14)
    fila_enc(NIT, COLOR_ENCABEZADO_FALTANTES, bold=False, size=10, italic=True)
    fila_enc(
        f"FACTURAS FALTANTES EN DIAN — {resultado.periodo_etiqueta.upper()}",
        COLOR_ENCABEZADO_FALTANTES, bold=True, size=12,
    )
    fila_enc(
        f"Período: {resultado.periodo_inicio} – {resultado.periodo_fin}  |  "
        f"Generado: {resultado.timestamp_ejecucion}",
        COLOR_AZUL_MEDIO, bold=False, size=10, italic=True,
    )
    return fila
