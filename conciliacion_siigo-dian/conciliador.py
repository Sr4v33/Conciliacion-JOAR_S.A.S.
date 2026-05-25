"""
Lógica de conciliación DIAN–Siigo
Rectificadora JOAR S.A.S.

Analiza qué facturas presentes en Siigo NO están registradas en DIAN.
"""

from __future__ import annotations

import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from dataclasses import dataclass, field
from typing import Any

from utils import (
    logger,
    normalizar_folio,
    extraer_folio_siigo,
    a_fecha,
    calcular_dias_antiguedad,
    prioridad_por_antiguedad,
    inferir_periodo,
)


# ---------------------------------------------------------------------------
# Colores de marcado en los archivos originales
# ---------------------------------------------------------------------------

FILL_AMARILLO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_ROJO = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


# ---------------------------------------------------------------------------
# Estructuras de datos
# ---------------------------------------------------------------------------

@dataclass
class RegistroFaltante:
    """Representa una fila de Siigo sin correspondencia en DIAN."""
    comprobante: str
    fecha_elaboracion: str
    identificacion: str
    razon_social: str
    factura_proveedor: str
    base_gravable: float
    base_exenta: float
    iva: float
    total: float
    # Campos enriquecidos durante la conciliación
    dias_antiguedad: int = 0
    prioridad: str = "Media"
    fecha_recepcion_dian: str = ""


@dataclass
class ResultadoConciliacion:
    """Resultado completo de la conciliación."""
    total_dian: int = 0
    total_siigo: int = 0
    total_conciliadas: int = 0
    total_faltantes: int = 0
    valor_faltantes: float = 0.0
    periodo_etiqueta: str = ""
    periodo_inicio: str = ""
    periodo_fin: str = ""
    timestamp_ejecucion: str = ""
    faltantes: list[RegistroFaltante] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Carga de archivos
# ---------------------------------------------------------------------------

def cargar_dian(ruta: str, tipos_doc: list[str], grupo: str) -> pd.DataFrame:
    """
    Carga el archivo DIAN, aplica los filtros definidos y retorna el DataFrame filtrado.
    Encabezados en fila 1 (header=0).
    """
    logger.info("Cargando archivo DIAN: %s", ruta)
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"Archivo DIAN no encontrado: {ruta}")

    df = pd.read_excel(ruta, header=0, dtype=str)

    # 1. Guardar la fila original de Excel ANTES de filtrar 
    # (índice 0 base pandas + 2 = fila real Excel tomando en cuenta el encabezado)
    df["_fila_excel"] = df.index + 2

    # Renombrar columnas por posición para garantizar acceso independiente del nombre exacto
    col_map = _mapear_columnas_dian(df)
    df = df.rename(columns=col_map)

    # Normalizar texto en columnas de filtro
    df["_tipo_doc"] = df["_tipo_doc"].fillna("").str.strip()
    df["_grupo"] = df["_grupo"].fillna("").str.strip()

    # Aplicar filtros
    mask = (df["_tipo_doc"].isin(tipos_doc)) & (df["_grupo"] == grupo)
    df_filtrado = df[mask].copy().reset_index(drop=True)

    # Normalizar folio DIAN a string
    df_filtrado["_folio_norm"] = df_filtrado["_folio"].apply(normalizar_folio)

    logger.info(
        "DIAN: %d filas totales → %d filas tras filtrar por tipo/grupo",
        len(df),
        len(df_filtrado),
    )
    return df_filtrado


def _mapear_columnas_dian(df: pd.DataFrame) -> dict[str, str]:
    """
    Mapea columnas del DataFrame DIAN a nombres internos según su posición de letra.
    A=0, B=1, C=2, D=3, I=8, J=9, K=10, N=13, AD=29, AE=30, AF=31
    """
    ncols = len(df.columns)
    mapeo = {}
    posiciones = {
        0: "_tipo_doc",   # A - Tipo de documento
        1: "_cufe",       # B - CUFE/CUDE
        2: "_folio",      # C - Folio
        3: "_prefijo",    # D - Prefijo
        8: "_fecha_recepcion",  # I - Fecha Recepción
        9: "_nit_emisor", # J - NIT Emisor
        10: "_nombre_emisor",  # K - Nombre Emisor
        13: "_iva",       # N - IVA
        29: "_total",     # AD - Total
        30: "_estado",    # AE - Estado
        31: "_grupo",     # AF - Grupo
    }
    for pos, nombre in posiciones.items():
        if pos < ncols:
            col_original = df.columns[pos]
            mapeo[col_original] = nombre
        else:
            logger.warning("Columna en posición %d no encontrada en DIAN (ncols=%d)", pos, ncols)
    return mapeo


def cargar_siigo(ruta: str) -> pd.DataFrame:
    """
    Carga el archivo Siigo. Encabezados en fila 8 (header=7), datos desde fila 9.
    """
    logger.info("Cargando archivo Siigo: %s", ruta)
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"Archivo Siigo no encontrado: {ruta}")

    df = pd.read_excel(ruta, header=7, dtype=str)

    # Guardar fila original de Excel (índice 0 base pandas + 9 = fila 9 en Excel)
    df["_fila_excel"] = df.index + 9

    col_map = _mapear_columnas_siigo(df)
    df = df.rename(columns=col_map)

    # Normalizar comprobante para el filtro
    df["_comprobante"] = df["_comprobante"].fillna("").astype(str).str.strip()

    # Omitir filas vacías y las de totales/pie de página (Total general, Procesado en...)
    mask_validas = (
        (df["_comprobante"] != "") &
        (df["_comprobante"].str.lower() != "nan") &
        (~df["_comprobante"].str.lower().str.contains("total", na=False)) &
        (~df["_comprobante"].str.lower().str.contains("procesado", na=False))
    )
    df = df[mask_validas].copy().reset_index(drop=True)

    logger.info("Siigo: %d registros cargados válidos", len(df))
    return df


def _mapear_columnas_siigo(df: pd.DataFrame) -> dict[str, str]:
    """
    A=0 Comprobante, B=1 Fecha elaboración, C=2 Identificación,
    D=3 Razón Social, E=4 Factura proveedor, F=5 Base gravable,
    G=6 Base exenta, H=7 IVA, I=8 Total
    """
    ncols = len(df.columns)
    posiciones = {
        0: "_comprobante",
        1: "_fecha_elab",
        2: "_identificacion",
        3: "_razon_social",
        4: "_factura_proveedor",
        5: "_base_gravable",
        6: "_base_exenta",
        7: "_iva",
        8: "_total",
    }
    mapeo = {}
    for pos, nombre in posiciones.items():
        if pos < ncols:
            mapeo[df.columns[pos]] = nombre
        else:
            logger.warning("Columna en posición %d no encontrada en Siigo (ncols=%d)", pos, ncols)
    return mapeo


# ---------------------------------------------------------------------------
# Proceso de conciliación
# ---------------------------------------------------------------------------

def conciliar(
    ruta_dian: str,
    ruta_siigo: str,
    tipos_doc: list[str],
    grupo: str,
) -> tuple[ResultadoConciliacion, str, str]:
    """
    Ejecuta el proceso completo de conciliación.

    Retorna:
        resultado: ResultadoConciliacion con métricas y lista de faltantes
        ruta_dian_marcado: ruta del archivo DIAN con celdas coloreadas
        ruta_siigo_marcado: ruta del archivo Siigo con celdas coloreadas
    """
    from datetime import datetime as dt

    # 1. Cargar archivos
    df_dian = cargar_dian(ruta_dian, tipos_doc, grupo)
    df_siigo = cargar_siigo(ruta_siigo)

    # 2. Construir set de folios DIAN (string normalizado → lista de filas Excel reales)
    dian_folios: dict[str, list[int]] = {}
    for idx, row in df_dian.iterrows():
        folio = normalizar_folio(row.get("_folio_norm", row.get("_folio", "")))
        fila_excel = int(row["_fila_excel"])
        dian_folios.setdefault(folio, []).append(fila_excel)

    # 3. Matching
    resultado = ResultadoConciliacion(
        total_dian=len(df_dian),
        total_siigo=len(df_siigo),
        timestamp_ejecucion=dt.now().strftime("%d/%m/%Y %H:%M:%S"),
    )

    dian_filas_match: set[int] = set()
    siigo_filas_conciliadas: list[int] = []
    siigo_filas_faltantes: list[int] = []

    for idx_s, row_s in df_siigo.iterrows():
        fila_excel_siigo = int(row_s["_fila_excel"])
        factura_prov = row_s.get("_factura_proveedor", "")
        folio_siigo = extraer_folio_siigo(factura_prov)
        folio_norm = normalizar_folio(folio_siigo)

        if folio_norm in dian_folios and folio_norm != "":
            # Match encontrado
            siigo_filas_conciliadas.append(fila_excel_siigo)
            for fila_dian in dian_folios[folio_norm]:
                dian_filas_match.add(fila_dian)
            resultado.total_conciliadas += 1
        else:
            # Sin match
            siigo_filas_faltantes.append(fila_excel_siigo)
            resultado.total_faltantes += 1

            total_val = _a_float(row_s.get("_total", 0))
            resultado.valor_faltantes += total_val

            fecha_elab = row_s.get("_fecha_elab", "")
            dias = calcular_dias_antiguedad(fecha_elab)
            prioridad = prioridad_por_antiguedad(dias)

            faltante = RegistroFaltante(
                comprobante=str(row_s.get("_comprobante", "") or ""),
                fecha_elaboracion=str(row_s.get("_fecha_elab", "") or ""),
                identificacion=str(row_s.get("_identificacion", "") or ""),
                razon_social=str(row_s.get("_razon_social", "") or ""),
                factura_proveedor=str(factura_prov or ""),
                base_gravable=_a_float(row_s.get("_base_gravable", 0)),
                base_exenta=_a_float(row_s.get("_base_exenta", 0)),
                iva=_a_float(row_s.get("_iva", 0)),
                total=total_val,
                dias_antiguedad=dias,
                prioridad=prioridad,
            )
            resultado.faltantes.append(faltante)

    # 4. Inferir período desde fechas Siigo
    fechas_siigo = df_siigo["_fecha_elab"].tolist() if "_fecha_elab" in df_siigo.columns else []
    etiqueta, f_inicio, f_fin = inferir_periodo(fechas_siigo)
    resultado.periodo_etiqueta = etiqueta
    resultado.periodo_inicio = f_inicio
    resultado.periodo_fin = f_fin

    logger.info(
        "Conciliación: %d conciliadas | %d faltantes | Valor faltantes: $ %s",
        resultado.total_conciliadas,
        resultado.total_faltantes,
        f"{resultado.valor_faltantes:,.0f}",
    )

    # 5. Colorear archivos originales
    ruta_dian_col = _colorear_dian(ruta_dian, dian_filas_match)
    ruta_siigo_col = _colorear_siigo(ruta_siigo, siigo_filas_conciliadas, siigo_filas_faltantes)

    return resultado, ruta_dian_col, ruta_siigo_col


# ---------------------------------------------------------------------------
# Coloreado de archivos originales
# ---------------------------------------------------------------------------

def _colorear_dian(ruta_original: str, filas_match: set[int]) -> str:
    """
    Abre el archivo DIAN original y marca en amarillo SÓLO la columna 'Folio'
    de las filas que tuvieron match.
    """
    ruta_salida = ruta_original.replace(".xlsx", "_marcado.xlsx")
    wb = openpyxl.load_workbook(ruta_original)
    ws = wb.active

    # Columna 'Folio' en la posición C del Excel (Columna 3)
    COLUMNA_FOLIO = 3

    for fila_excel in filas_match:
        ws.cell(row=fila_excel, column=COLUMNA_FOLIO).fill = FILL_AMARILLO

    wb.save(ruta_salida)
    logger.info("Archivo DIAN marcado guardado: %s", ruta_salida)
    return ruta_salida


def _colorear_siigo(
    ruta_original: str,
    filas_conciliadas: list[int],
    filas_faltantes: list[int],
) -> str:
    """
    Abre el archivo Siigo original y marca SÓLO la columna 'Factura Proveedor'.
    - Amarillo: conciliadas
    - Rojo: faltantes
    """
    ruta_salida = ruta_original.replace(".xlsx", "_marcado.xlsx")
    wb = openpyxl.load_workbook(ruta_original)
    ws = wb.active

    # Columna 'Factura proveedor' en la posición E del Excel (Columna 5)
    COLUMNA_FACTURA = 5

    for fila_excel in filas_conciliadas:
        ws.cell(row=fila_excel, column=COLUMNA_FACTURA).fill = FILL_AMARILLO

    for fila_excel in filas_faltantes:
        ws.cell(row=fila_excel, column=COLUMNA_FACTURA).fill = FILL_ROJO

    wb.save(ruta_salida)
    logger.info("Archivo Siigo marcado guardado: %s", ruta_salida)
    return ruta_salida


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _a_float(valor: Any) -> float:
    """Convierte un valor a float de forma segura."""
    try:
        if valor is None or str(valor).strip() in ("", "nan", "None"):
            return 0.0
        return float(str(valor).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0