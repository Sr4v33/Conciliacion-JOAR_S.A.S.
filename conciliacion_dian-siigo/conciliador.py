"""
Lógica de conciliación inversa DIAN→Siigo
Rectificadora JOAR S.A.S.

Analiza qué facturas presentes en DIAN NO están registradas en Siigo.
"""

from __future__ import annotations

import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from dataclasses import dataclass, field
from typing import Any

from utils import (
    logger,
    normalizar_folio,
    extraer_folio_siigo,
    a_fecha,
    calcular_dias_antiguedad,
    prioridad_por_antiguedad,
    inferir_periodo,
)


# ---------------------------------------------------------------------------
# Colores de marcado en los archivos originales
# ---------------------------------------------------------------------------

FILL_AMARILLO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_ROJO     = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


# ---------------------------------------------------------------------------
# Estructuras de datos
# ---------------------------------------------------------------------------

@dataclass
class RegistroFaltanteDIAN:
    """Representa una fila de DIAN sin correspondencia en Siigo."""
    tipo_documento: str
    cufe: str
    folio: str
    prefijo: str
    fecha_recepcion: str
    nit_emisor: str
    nombre_emisor: str
    iva: float
    total: float
    estado: str
    grupo: str
    # Campos enriquecidos durante la conciliación
    dias_antiguedad: int = 0
    prioridad: str = "Media"


@dataclass
class ResultadoConciliacionInversa:
    """Resultado completo de la conciliación inversa."""
    total_dian: int = 0
    total_siigo: int = 0
    total_conciliadas: int = 0
    total_faltantes: int = 0
    valor_faltantes: float = 0.0
    periodo_etiqueta: str = ""
    periodo_inicio: str = ""
    periodo_fin: str = ""
    timestamp_ejecucion: str = ""
    faltantes: list[RegistroFaltanteDIAN] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Carga de archivos  (reutiliza exactamente la misma lógica que el original)
# ---------------------------------------------------------------------------

def cargar_dian(ruta: str, tipos_doc: list[str], grupo: str) -> pd.DataFrame:
    """
    Carga el archivo DIAN, aplica los filtros definidos y retorna el DataFrame filtrado.
    Encabezados en fila 1 (header=0).
    """
    logger.info("Cargando archivo DIAN: %s", ruta)
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"Archivo DIAN no encontrado: {ruta}")

    df = pd.read_excel(ruta, header=0, dtype=str)
    df["_fila_excel"] = df.index + 2

    col_map = _mapear_columnas_dian(df)
    df = df.rename(columns=col_map)

    df["_tipo_doc"] = df["_tipo_doc"].fillna("").str.strip()
    df["_grupo"]    = df["_grupo"].fillna("").str.strip()

    mask = (df["_tipo_doc"].isin(tipos_doc)) & (df["_grupo"] == grupo)
    df_filtrado = df[mask].copy().reset_index(drop=True)
    df_filtrado["_folio_norm"] = df_filtrado["_folio"].apply(normalizar_folio)

    logger.info(
        "DIAN: %d filas totales → %d filas tras filtrar por tipo/grupo",
        len(df),
        len(df_filtrado),
    )
    return df_filtrado


def _mapear_columnas_dian(df: pd.DataFrame) -> dict[str, str]:
    """
    A=0 Tipo doc, B=1 CUFE, C=2 Folio, D=3 Prefijo,
    I=8 Fecha recepción, J=9 NIT emisor, K=10 Nombre emisor,
    N=13 IVA, AD=29 Total, AE=30 Estado, AF=31 Grupo
    """
    ncols = len(df.columns)
    posiciones = {
        0:  "_tipo_doc",
        1:  "_cufe",
        2:  "_folio",
        3:  "_prefijo",
        8:  "_fecha_recepcion",
        9:  "_nit_emisor",
        10: "_nombre_emisor",
        13: "_iva",
        29: "_total",
        30: "_estado",
        31: "_grupo",
    }
    mapeo = {}
    for pos, nombre in posiciones.items():
        if pos < ncols:
            mapeo[df.columns[pos]] = nombre
        else:
            logger.warning(
                "Columna en posición %d no encontrada en DIAN (ncols=%d)", pos, ncols
            )
    return mapeo


def cargar_siigo(ruta: str) -> pd.DataFrame:
    """
    Carga el archivo Siigo. Encabezados en fila 8 (header=7), datos desde fila 9.
    """
    logger.info("Cargando archivo Siigo: %s", ruta)
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"Archivo Siigo no encontrado: {ruta}")

    df = pd.read_excel(ruta, header=7, dtype=str)
    df["_fila_excel"] = df.index + 9

    col_map = _mapear_columnas_siigo(df)
    df = df.rename(columns=col_map)

    df["_comprobante"] = df["_comprobante"].fillna("").astype(str).str.strip()

    mask_validas = (
        (df["_comprobante"] != "") &
        (df["_comprobante"].str.lower() != "nan") &
        (~df["_comprobante"].str.lower().str.contains("total",     na=False)) &
        (~df["_comprobante"].str.lower().str.contains("procesado", na=False))
    )
    df = df[mask_validas].copy().reset_index(drop=True)

    logger.info("Siigo: %d registros cargados válidos", len(df))
    return df


def _mapear_columnas_siigo(df: pd.DataFrame) -> dict[str, str]:
    ncols = len(df.columns)
    posiciones = {
        0: "_comprobante",
        1: "_fecha_elab",
        2: "_identificacion",
        3: "_razon_social",
        4: "_factura_proveedor",
        5: "_base_gravable",
        6: "_base_exenta",
        7: "_iva",
        8: "_total",
    }
    mapeo = {}
    for pos, nombre in posiciones.items():
        if pos < ncols:
            mapeo[df.columns[pos]] = nombre
        else:
            logger.warning(
                "Columna en posición %d no encontrada en Siigo (ncols=%d)", pos, ncols
            )
    return mapeo


# ---------------------------------------------------------------------------
# Proceso de conciliación inversa
# ---------------------------------------------------------------------------

def conciliar_inverso(
    ruta_dian: str,
    ruta_siigo: str,
    tipos_doc: list[str],
    grupo: str,
) -> tuple[ResultadoConciliacionInversa, str, str]:
    """
    Ejecuta la conciliación inversa: detecta facturas que están en DIAN
    pero NO tienen registro en Siigo.

    Retorna:
        resultado     : ResultadoConciliacionInversa con métricas y lista de faltantes
        ruta_dian_marcado  : ruta del archivo DIAN con celdas coloreadas
        ruta_siigo_marcado : ruta del archivo Siigo con celdas coloreadas
    """
    from datetime import datetime as dt

    # 1. Cargar archivos
    df_dian  = cargar_dian(ruta_dian, tipos_doc, grupo)
    df_siigo = cargar_siigo(ruta_siigo)

    # 2. Construir set de folios Siigo (normalizados)
    siigo_folios: set[str] = set()
    siigo_folio_a_filas: dict[str, list[int]] = {}
    for idx, row in df_siigo.iterrows():
        fp    = row.get("_factura_proveedor", "")
        folio = normalizar_folio(extraer_folio_siigo(fp))
        if folio:
            siigo_folios.add(folio)
            siigo_folio_a_filas.setdefault(folio, []).append(int(row["_fila_excel"]))

    # 3. Matching (iteramos sobre DIAN, buscamos en Siigo)
    resultado = ResultadoConciliacionInversa(
        total_dian=len(df_dian),
        total_siigo=len(df_siigo),
        timestamp_ejecucion=dt.now().strftime("%d/%m/%Y %H:%M:%S"),
    )

    dian_filas_conciliadas: list[int] = []
    dian_filas_faltantes:   list[int] = []
    siigo_filas_match:      set[int]  = set()

    for idx_d, row_d in df_dian.iterrows():
        fila_excel_dian = int(row_d["_fila_excel"])
        folio_dian      = normalizar_folio(row_d.get("_folio_norm", row_d.get("_folio", "")))

        if folio_dian and folio_dian in siigo_folios:
            # Match encontrado
            dian_filas_conciliadas.append(fila_excel_dian)
            for fila_siigo in siigo_folio_a_filas.get(folio_dian, []):
                siigo_filas_match.add(fila_siigo)
            resultado.total_conciliadas += 1
        else:
            # Sin match en Siigo
            dian_filas_faltantes.append(fila_excel_dian)
            resultado.total_faltantes += 1

            total_val = _a_float(row_d.get("_total", 0))
            resultado.valor_faltantes += total_val

            fecha_rec = str(row_d.get("_fecha_recepcion", "") or "")
            dias      = calcular_dias_antiguedad(fecha_rec)
            prioridad = prioridad_por_antiguedad(dias)

            faltante = RegistroFaltanteDIAN(
                tipo_documento=str(row_d.get("_tipo_doc",       "") or ""),
                cufe          =str(row_d.get("_cufe",           "") or ""),
                folio         =str(row_d.get("_folio",          "") or ""),
                prefijo       =str(row_d.get("_prefijo",        "") or ""),
                fecha_recepcion=fecha_rec,
                nit_emisor    =str(row_d.get("_nit_emisor",     "") or ""),
                nombre_emisor =str(row_d.get("_nombre_emisor",  "") or ""),
                iva           =_a_float(row_d.get("_iva",   0)),
                total         =total_val,
                estado        =str(row_d.get("_estado",         "") or ""),
                grupo         =str(row_d.get("_grupo",          "") or ""),
                dias_antiguedad=dias,
                prioridad     =prioridad,
            )
            resultado.faltantes.append(faltante)

    # 4. Inferir período desde fechas de recepción DIAN
    fechas_dian = df_dian["_fecha_recepcion"].tolist() if "_fecha_recepcion" in df_dian.columns else []
    etiqueta, f_inicio, f_fin = inferir_periodo(fechas_dian)
    resultado.periodo_etiqueta = etiqueta
    resultado.periodo_inicio   = f_inicio
    resultado.periodo_fin      = f_fin

    logger.info(
        "Conciliación inversa: %d conciliadas | %d faltantes en Siigo | Valor: $ %s",
        resultado.total_conciliadas,
        resultado.total_faltantes,
        f"{resultado.valor_faltantes:,.0f}",
    )

    # 5. Colorear archivos originales
    ruta_dian_col  = _colorear_dian(ruta_dian,  dian_filas_conciliadas, dian_filas_faltantes)
    ruta_siigo_col = _colorear_siigo(ruta_siigo, siigo_filas_match)

    return resultado, ruta_dian_col, ruta_siigo_col


# ---------------------------------------------------------------------------
# Coloreado de archivos originales
# ---------------------------------------------------------------------------

def _colorear_dian(
    ruta_original: str,
    filas_conciliadas: list[int],
    filas_faltantes: list[int],
) -> str:
    """
    Marca en la columna 'Folio' (C):
    - Amarillo: conciliadas (tienen match en Siigo)
    - Rojo    : faltantes   (no tienen match en Siigo)
    """
    ruta_salida = ruta_original.replace(".xlsx", "_marcado.xlsx")
    wb = openpyxl.load_workbook(ruta_original)
    ws = wb.active

    COLUMNA_FOLIO = 3   # Columna C

    for fila_excel in filas_conciliadas:
        ws.cell(row=fila_excel, column=COLUMNA_FOLIO).fill = FILL_AMARILLO

    for fila_excel in filas_faltantes:
        ws.cell(row=fila_excel, column=COLUMNA_FOLIO).fill = FILL_ROJO

    wb.save(ruta_salida)
    logger.info("Archivo DIAN marcado guardado: %s", ruta_salida)
    return ruta_salida


def _colorear_siigo(ruta_original: str, filas_match: set[int]) -> str:
    """
    Marca en amarillo la columna 'Factura Proveedor' (E) de las filas con match.
    Las filas sin match no se tocan (no son faltantes desde la perspectiva Siigo).
    """
    ruta_salida = ruta_original.replace(".xlsx", "_marcado.xlsx")
    wb = openpyxl.load_workbook(ruta_original)
    ws = wb.active

    COLUMNA_FACTURA = 5   # Columna E

    for fila_excel in filas_match:
        ws.cell(row=fila_excel, column=COLUMNA_FACTURA).fill = FILL_AMARILLO

    wb.save(ruta_salida)
    logger.info("Archivo Siigo marcado guardado: %s", ruta_salida)
    return ruta_salida


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _a_float(valor: Any) -> float:
    try:
        if valor is None or str(valor).strip() in ("", "nan", "None"):
            return 0.0
        return float(str(valor).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0